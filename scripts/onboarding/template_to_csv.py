#!/usr/bin/env python3
"""Convert a filled EnergyLens intake template (.xlsx) into the 4 bronze source CSVs.

Usage:
    python template_to_csv.py <filled_template.xlsx> [output_dir]

Reads the Buildings / Energy readings / Solar / Battery sheets (row 1 = headers,
row 2 = EXAMPLE, row 3+ = data) and writes the EXACT filenames the bronze ingestion
notebook expects in Files/sample-data/:
    building_master.csv · raw_energy_readings.csv · raw_solar_generation.csv · raw_battery_status.csv

It validates required fields and that every energy/solar/battery building_id exists in
Buildings, skips the EXAMPLE row and blanks, and only writes Solar/Battery CSVs when those
sheets contain data. Exits non-zero (and writes nothing partial-but-broken) on hard errors.
"""
import csv
import sys
from pathlib import Path

from openpyxl import load_workbook

# sheet -> (output csv, required columns that must be non-empty)
SHEETS = {
    "Buildings": (
        "building_master.csv",
        ["building_id", "building_name", "country_code", "city",
         "gross_floor_area_m2", "conditioned_area_m2", "year_built",
         "building_type", "primary_hvac_system"],
    ),
    "Energy readings": (
        "raw_energy_readings.csv",
        ["building_id", "timestamp_utc", "raw_value", "raw_unit"],
    ),
    "Solar (optional)": (
        "raw_solar_generation.csv",
        ["building_id", "timestamp_utc", "generated_raw", "exported_raw", "raw_unit"],
    ),
    "Battery (optional)": (
        "raw_battery_status.csv",
        ["building_id", "timestamp_utc", "soc_raw"],
    ),
}


def read_sheet(ws):
    headers = [c.value for c in ws[1]]
    while headers and headers[-1] in (None, ""):
        headers.pop()
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        d = {headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))}
        bid = "" if d.get("building_id") is None else str(d["building_id"]).strip()
        if not bid or bid.upper() == "EXAMPLE":
            continue
        if all(v in (None, "") for v in d.values()):
            continue
        rows.append(d)
    return headers, rows


def main():
    if len(sys.argv) < 2:
        print("usage: template_to_csv.py <filled.xlsx> [out_dir]")
        sys.exit(2)
    src = Path(sys.argv[1])
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else src.parent
    out.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(src, data_only=True)
    errors, warnings, pending = [], [], []
    building_ids = set()

    for sheet, (csv_name, required) in SHEETS.items():
        if sheet not in wb.sheetnames:
            warnings.append(f"sheet '{sheet}' missing — skipped")
            continue
        headers, rows = read_sheet(wb[sheet])
        if sheet == "Buildings":
            building_ids = {str(r["building_id"]).strip() for r in rows}
        if not rows:
            if sheet in ("Buildings", "Energy readings"):
                errors.append(f"{sheet}: no data rows (this sheet is required)")
            else:
                warnings.append(f"{sheet}: empty — CSV not written")
            continue
        for i, r in enumerate(rows, start=3):
            for f in required:
                if r.get(f) in (None, ""):
                    errors.append(f"{sheet} row {i}: missing required '{f}'")
            if sheet != "Buildings":
                bid = str(r.get("building_id")).strip()
                if bid not in building_ids:
                    errors.append(
                        f"{sheet} row {i}: building_id '{bid}' is not in the Buildings sheet"
                    )
        pending.append((csv_name, headers, rows))

    if errors:
        print("=== template_to_csv ===")
        for w in warnings:
            print(f"  WARN: {w}")
        for e in errors:
            print(f"  ERROR: {e}")
        print(f"FAILED with {len(errors)} error(s) — fix the template and re-run. No CSVs written.")
        sys.exit(1)

    written = []
    for csv_name, headers, rows in pending:
        with open(out / csv_name, "w", newline="", encoding="utf-8") as fh:
            w = csv.DictWriter(fh, fieldnames=headers)
            w.writeheader()
            for r in rows:
                w.writerow({h: ("" if r.get(h) is None else r.get(h)) for h in headers})
        written.append((csv_name, len(rows)))

    print("=== template_to_csv ===")
    for n, c in written:
        print(f"  wrote {n}: {c} rows")
    for w in warnings:
        print(f"  note: {w}")
    print(f"OK — {len(written)} CSV(s) ready. Upload them to the lakehouse Files/sample-data/ (Step 3).")


if __name__ == "__main__":
    main()
